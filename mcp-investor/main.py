from fastapi import FastAPI, Body
from fastapi_mcp import FastApiMCP
import pandas as pd
from openpyxl import load_workbook
from pydantic import BaseModel
import uvicorn
import math

SPREADSHEET_FILEPATH = "investments.xlsx"

app = FastAPI()


@app.get("/")
async def root():
    return {"message": "MCP Investments Server"}


@app.get(
    "/list_availables_sheets",
    operation_id="list_availables_sheets"
)
async def list_availables_sheets():
    """
    Read the excel file containing the investment information,
    and return a list of available sheets
    """
    xl = pd.ExcelFile(SPREADSHEET_FILEPATH)
    return {"message": xl.sheet_names}


@app.get(
    "/update_investment_distributions",
    operation_id="update_investment_distributions"
)
def update_investment_distributions() -> None:
    """
    Update the investment distributions in the "Distribution" sheet.
    """
    wb = load_workbook(SPREADSHEET_FILEPATH)
    total_values = []
    for sheet in wb.sheetnames[1:]:
        df = pd.read_excel(SPREADSHEET_FILEPATH, sheet_name=sheet)
        invest_sum = round(float(df.iloc[:-1, -1].sum()), 2)
        total_values.append(invest_sum)
        for i in range(1, len(wb["Distribution"]["A"])):
            if wb["Distribution"]["A"][i].value == sheet:
                wb["Distribution"]["D"][i].value = invest_sum
    wb.save(SPREADSHEET_FILEPATH)


@app.get(
    "/get_investment_distribution",
    operation_id="get_investment_distribution"
)
def get_investment_distribution(sheet_name: str, total_value: int) -> dict:
    """
    Given a total investment value, this function returns the 
    monetary value for each investment based on each investment percentage.
    """
    df = pd.read_excel("investments.xlsx", sheet_name=sheet_name).fillna(0)
    def_col = df.columns[0]
    distribution = {}
    sum_to_be_invested = 0
    new_value = total_value
    while math.ceil(sum_to_be_invested) != total_value:
        for _, row in df.iterrows():
            if row["Target Percentage"] >= row["Current Percentage"] and \
                row["Target Percentage"] < 1.0:
                if distribution.get(row[def_col]) is None:
                    distribution[row[def_col]] = row["Target Percentage"] * new_value
                else:
                    distribution[row[def_col]] += row["Target Percentage"] * new_value
        sum_to_be_invested = sum(list(distribution.values()))
        new_value = total_value - sum_to_be_invested
    for key in distribution.keys():
        distribution[key] = round(distribution[key], 2)
    return distribution


class InvestmentRequest(BaseModel):
    sheet_name: str
    distribution_dict: dict

@app.post(
    "/execute_investments",
    operation_id="execute_investments"
)
async def execute_investments(
    request: InvestmentRequest
) -> None:
    """
    Updates the investment values in the given sheet
    based on the distribution dictionary.
    Use the return value of the get_investment_distribution
    function as the parameter distribution_dict.
    """
    wb = load_workbook(SPREADSHEET_FILEPATH)
    ws = wb[request.sheet_name]
    for cols in list(ws.rows)[0]:
        if cols.value == "Investment":
            investment_letter = cols.column_letter
    for idx, tick_xls in enumerate(ws["A"][1:], start=1):
        current_investment = ws[investment_letter][idx].value or 0
        for tick_inv, quant in request.distribution_dict.items():
            if tick_xls.value == tick_inv and not isinstance(current_investment, str):
                ws[investment_letter][idx].value = current_investment + quant
    wb.save(SPREADSHEET_FILEPATH)


mcp = FastApiMCP(
    app,
    name="MCP Investments Server",
    description="Simple MCP Investments Server"
)

mcp.mount()

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
